#!/usr/bin/env python3
"""Scrapes Beckett.com checklist .xlsx files into src/db/checklist-data/*.json
(the checklist_cards import format — see src/db/checklistImport.ts).

Two modes:
  Single product page:
    python3 scrape_beckett_checklist.py --url https://www.beckett.com/news/2025-topps-chrome-football-cards/

  Whole yearly archive (finds every product link on the archive page and
  scrapes each one, skipping pages with no checklist .xlsx):
    python3 scrape_beckett_checklist.py --archive https://www.beckett.com/news/2025-football-cards-release-dates-checklists-and-set-information/

Only football (NFL) is supported right now (sportEnum has no CFL value —
CFL product pages are detected and skipped, not miscategorized as NFL).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path
from urllib.parse import urljoin

import requests
import openpyxl

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
HEADERS = {"User-Agent": UA}

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "src" / "db" / "checklist-data"

MANUFACTURERS = [
    "Topps", "Panini", "Donruss", "Leaf", "Upper Deck", "Bowman", "Score",
    "Pro Set", "Press Pass",
]

TEAM_SHEET_CANDIDATES = ["Teams", "Team Sets"]

EXCLUDE_SLUG_PATTERNS = [
    "release-dates-checklists", "-and-set-information", "category/", "author/",
    "box-busters", "rookie-cards", "parallel-guide", "image-gallery",
    "countdown-calendar", "complete-factory-set", "box-set", "top-selling",
    "hot-cold", "checklist-and-details",  # e.g. "topps-now-nfl-draft-checklist-and-details" — not a product xlsx page
]


def fetch(url: str) -> str:
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return resp.text


def find_product_links(archive_html: str, archive_url: str, year: str) -> list[str]:
    hrefs = re.findall(r'href="(https://www\.beckett\.com/news/[^"]*)"', archive_html)
    seen = set()
    out = []
    for h in hrefs:
        h = h.rstrip("/") + "/"
        if h in seen:
            continue
        seen.add(h)
        slug = h.lower()
        if "football" not in slug:
            continue
        if year not in slug:
            continue
        if any(p in slug for p in EXCLUDE_SLUG_PATTERNS):
            continue
        if "cfl" in slug:
            continue  # Canadian Football — not a supported sport enum value
        out.append(h)
    return out


def find_xlsx_url(page_html: str, page_url: str) -> str | None:
    m = re.search(r'href="([^"]*\.xlsx[^"]*)"', page_html)
    if not m:
        return None
    return urljoin(page_url, m.group(1))


def product_name_from_filename(xlsx_url: str) -> str:
    fname = xlsx_url.rsplit("/", 1)[-1]
    fname = re.sub(r"\.xlsx$", "", fname, flags=re.IGNORECASE)
    fname = re.sub(r"[-_]Checklist$", "", fname, flags=re.IGNORECASE)
    return fname.replace("-", " ").strip()


def guess_manufacturer(product_name: str) -> str | None:
    for m in MANUFACTURERS:
        if product_name.lower().startswith(m.lower()) or f" {m.lower()} " in f" {product_name.lower()} ":
            return m
    return None


def guess_year(product_name: str) -> str:
    m = re.match(r"^(\d{4})", product_name.strip())
    return m.group(1) if m else "unknown"


def _row_dict(product, subset, card_number, player, team, rc_flag, sport, year, manufacturer, source_url):
    player_clean = str(player).rstrip(",").strip()
    card_number_clean = "" if card_number is None else str(card_number).strip()
    subset_clean = str(subset).strip()
    team_clean = None if team is None else str(team).strip()
    is_rookie = bool(rc_flag) and str(rc_flag).strip().upper() == "RC"
    is_auto = "autograph" in subset_clean.lower()
    return {
        "sport": sport,
        "year": year,
        "manufacturer": manufacturer,
        "product": product,
        "subset": subset_clean,
        "card_number": card_number_clean,
        "player": player_clean,
        "team": team_clean,
        "parallel_name": None,
        "print_run": None,
        "is_auto": is_auto,
        "is_rookie": is_rookie,
        "source_url": source_url,
    }, (product, subset_clean, card_number_clean, player_clean)


def parse_team_sheet(ws, product, sport, year, manufacturer, source_url):
    rows = []
    seen = set()
    for row in ws.iter_rows(values_only=True):
        subset, card_number, player, team, rc_flag = (list(row) + [None] * 5)[:5]
        if not subset or player is None:
            continue
        rd, key = _row_dict(product, subset, card_number, player, team, rc_flag, sport, year, manufacturer, source_url)
        if key in seen:
            continue
        seen.add(key)
        rows.append(rd)
    return rows


def parse_full_checklist_sheet(ws, product, sport, year, manufacturer, source_url):
    """Fallback for products with no Teams/Team Sets sheet (e.g. Leaf brand
    checklists) — 'Full Checklist' is present on every product we've seen
    regardless of brand, but its column count varies (Topps: card_number,
    player, team, [rc]; Leaf: card_number, player only — no team). Section
    headers are single-non-null-cell rows naming the subset; a following
    "<N> cards" line is metadata, not data."""
    count_re = re.compile(r"^\d+\s+cards?$", re.IGNORECASE)
    rows = []
    seen = set()
    current_subset = None
    for row in ws.iter_rows(values_only=True):
        non_null = [c for c in row if c is not None]
        if len(non_null) <= 1:
            if non_null and not count_re.match(str(non_null[0]).strip()):
                current_subset = str(non_null[0]).strip()
            continue
        if current_subset is None or row[1] is None:
            continue
        card_number, player = row[0], row[1]
        team = row[2] if len(row) > 2 else None
        rc_flag = row[3] if len(row) > 3 else None
        rd, key = _row_dict(product, current_subset, card_number, player, team, rc_flag, sport, year, manufacturer, source_url)
        if key in seen:
            continue
        seen.add(key)
        rows.append(rd)
    return rows


MASTER_SHEET_CANDIDATES = ["Master Checklist", "Master", "Checklist"]


def _split_card_set(card_set: str, page_product: str, page_year: str) -> tuple[str, str, str]:
    """Splits a Master Checklist 'CARD SET' string into (year, product,
    subset). Most rows are just a subset/parallel name ('Base', 'Base
    Century Blue', 'College Materials Signature Booklet') implicitly
    belonging to *this page's* current release — the year/product is only
    written out when a row is a historical cross-year entry bundled into
    the same cumulative product-line sheet (e.g. '2017 Panini Flawless
    Collegiate - Flawless Team Slogan Signatures Black'), which needs its
    own year/product distinct from the page's."""
    m = re.match(r"^(\d{4})\s+(.*)$", card_set.strip())
    if not m:
        return page_year, page_product, card_set.strip()

    row_year, rest = m.group(1), m.group(2)
    if " - " in rest:
        product_part, subset_part = rest.split(" - ", 1)
    else:
        product_part, subset_part = rest, "Base"
    return row_year, product_part.strip(), subset_part.strip()


def parse_master_checklist_sheet(ws, page_product, sport, page_year, page_manufacturer, source_url):
    header = [str(c).strip().lower() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def find_col(*keywords):
        for i, h in enumerate(header):
            if any(k in h for k in keywords):
                return i
        return None

    col_set = find_col("card set", "set")
    col_num = find_col("card number", "number", "#")
    col_player = find_col("athlete", "player", "name")
    col_team = find_col("team")
    if col_set is None or col_num is None or col_player is None:
        return []

    rows = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_set >= len(row) or row[col_set] is None or col_player >= len(row) or row[col_player] is None:
            continue
        row_year, product_part, subset_part = _split_card_set(str(row[col_set]), page_product, page_year)
        manufacturer = guess_manufacturer(product_part) or page_manufacturer
        card_number = row[col_num] if col_num is not None and col_num < len(row) else None
        team = row[col_team] if col_team is not None and col_team < len(row) else None

        rd, key = _row_dict(
            product_part, subset_part, card_number, row[col_player], team, None,
            sport, row_year, manufacturer, source_url,
        )
        if key in seen:
            continue
        seen.add(key)
        rows.append(rd)
    return rows


def parse_workbook(xlsx_bytes: bytes, product: str, manufacturer: str, year: str, sport: str, source_url: str):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    sheet_name = next((s for s in TEAM_SHEET_CANDIDATES if s in wb.sheetnames), None)
    if sheet_name is not None:
        rows = parse_team_sheet(wb[sheet_name], product, sport, year, manufacturer, source_url)
        if rows:
            return rows, None

    if "Full Checklist" in wb.sheetnames:
        rows = parse_full_checklist_sheet(wb["Full Checklist"], product, sport, year, manufacturer, source_url)
        if rows:
            return rows, None

    master_sheet_name = next((s for s in MASTER_SHEET_CANDIDATES if s in wb.sheetnames), None)
    if master_sheet_name is not None:
        rows = parse_master_checklist_sheet(wb[master_sheet_name], product, sport, year, manufacturer, source_url)
        if rows:
            return rows, None

    return None, f"no usable sheet found (sheets: {wb.sheetnames})"


def slugify(product_name: str) -> str:
    s = product_name.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def scrape_product_page(page_url: str, sport: str = "NFL") -> dict:
    """Returns a result dict: {status: ok|skipped|error, ...}"""
    try:
        html = fetch(page_url)
    except Exception as e:
        return {"status": "error", "url": page_url, "reason": f"fetch failed: {e}"}

    xlsx_url = find_xlsx_url(html, page_url)
    if not xlsx_url:
        return {"status": "skipped", "url": page_url, "reason": "no .xlsx link on page"}

    product = product_name_from_filename(xlsx_url)
    manufacturer = guess_manufacturer(product) or "Unknown"
    year = guess_year(product)

    try:
        xlsx_resp = requests.get(xlsx_url, headers=HEADERS, timeout=60)
        xlsx_resp.raise_for_status()
    except Exception as e:
        return {"status": "error", "url": page_url, "reason": f"xlsx download failed: {e}"}

    try:
        rows, err = parse_workbook(xlsx_resp.content, product, manufacturer, year, sport, page_url)
    except Exception as e:
        return {"status": "error", "url": page_url, "reason": f"parse failed: {e}"}

    if err:
        return {"status": "skipped", "url": page_url, "product": product, "reason": err}

    out_path = OUT_DIR / f"{slugify(product)}.json"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(rows, f, indent=2)

    return {
        "status": "ok",
        "url": page_url,
        "product": product,
        "rows": len(rows),
        "out_path": str(out_path.relative_to(REPO_ROOT)),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", help="single product page URL")
    ap.add_argument("--archive", help="yearly archive page URL")
    ap.add_argument("--year", help="4-digit year (required with --archive, used to filter links)")
    ap.add_argument("--delay", type=float, default=1.0, help="seconds between requests (politeness)")
    args = ap.parse_args()

    results = []

    if args.url:
        results.append(scrape_product_page(args.url))
    elif args.archive:
        if not args.year:
            print("--year is required with --archive", file=sys.stderr)
            sys.exit(1)
        archive_html = fetch(args.archive)
        links = find_product_links(archive_html, args.archive, args.year)
        print(f"Found {len(links)} candidate product links for {args.year}", file=sys.stderr)
        for link in links:
            time.sleep(args.delay)
            result = scrape_product_page(link)
            results.append(result)
            print(f"  [{result['status']}] {link} — {result.get('reason', result.get('rows'))}", file=sys.stderr)
    else:
        print("Provide --url or --archive", file=sys.stderr)
        sys.exit(1)

    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main()
